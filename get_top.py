import time

import bs4
import requests
import re
import os
from openpyxl import Workbook, load_workbook


def get_movie_data_from_soup(soup: bs4.element.ResultSet):
    try:
        actors = []
        for i in soup.find_all(href=re.compile("/name/nm")):
            actors.append(i.text)
        return {
            "name": soup.h3.a.text,
            "year":soup.find("span", class_="lister-item-year text-muted unbold").text.strip()[1:-1],
            "type": soup.find("span", class_="genre").text.strip(),
            # "rating": soup.find("span", class_="ipl-rating-star__rating").text.strip(),
            "rating": soup.strong.text,
            "director": actors[0] if actors is not None else None,
            "actors": actors[1:] if actors is not None else None,
            "votes": soup.find_all("span")[-4].text.strip(),
            "page_link": f"https://www.imdb.com{soup.a.get('href')}"
        }
    except Exception as e:
        print(e)
        print(f"https://www.imdb.com{soup.a.get('href')}")


def get_imdb_top_movies(num_movies: int = 5) -> tuple:
    """Get the top num_movies most highly rated movies from IMDB and
    return a tuple of dicts describing each movie's name, genre, rating, and URL.

    Args:
        num_movies: The number of movies to get. Defaults to 5.

    Returns:
        A list of tuples containing information about the top n movies.

    >>> len(get_imdb_top_movies(5))
    5
    >>> len(get_imdb_top_movies(-3))
    0
    >>> len(get_imdb_top_movies(4.99999))
    4
    """
    num_movies = int(float(num_movies))
    if num_movies < 1:
        return ()
    # base_url = (
    #     "https://www.imdb.com/search/title?title_type="
    #     f"feature&sort=num_votes,desc&count={num_movies}"
    # )
    base_url = (
        "https://www.imdb.com/search/title/?title_type=feature&release_date=2020-01-11,"
        f"2023-06-20&countries=cn&languages=zh&sort=num_votes,desc&start={num_movies}"
    )
    source = bs4.BeautifulSoup(requests.get(base_url).content, "html.parser")
    return tuple(
        get_movie_data_from_soup(movie)
        # for movie in source.find_all("div", class_="lister-item mode-detail")
        for movie in source.find_all("div", class_="lister-item mode-advanced")
    )


if __name__ == "__main__":
    import json

    # num_movies = int(input("How many movies would you like to see? "))
    num_movies = 500
    # result = ", ".join(json.dumps(movie, indent=4) for movie in get_imdb_top_movies(num_movies))
    #
    # print(result)

    # with open("result2.txt", "w") as f:
    #     f.write(result)
    #
    # f.close()
    filepath = 'data.xlsx'

    for start in [1, 51, 101, 151, 201, 251, 301, 351, 401, 451, 501]:
        time.sleep(5)
        for movie in get_imdb_top_movies(start):
            try:
                allinfo = []
                allinfo.append(movie['name'])
                allinfo.append(movie['year'])
                allinfo.append(movie['type'])
                allinfo.append(movie['rating'])
                allinfo.append(movie['director'])
                allinfo.append(' '.join(movie['actors']))
                allinfo.append(movie['votes'])
                if not os.path.exists(filepath):
                    tableTitle = ['name','year','type','rating','director','actors','votes']
                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'sheet1'
                    ws.append(tableTitle)
                    wb.save('data.xlsx')
                wb = load_workbook(filepath)
                ws = wb.active
                ws.title = 'sheet1'
                ws.append(allinfo)
                wb.save(filepath)
            except Exception as e:
                print(e)
                print(allinfo)